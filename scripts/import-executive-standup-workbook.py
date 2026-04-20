#!/usr/bin/env python3
"""
Import the legacy standup workbook into exec_standup_* tables.

Usage:
  HAVEN_ORGANIZATION_ID=... NEXT_PUBLIC_SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... \
  python3 scripts/import-executive-standup-workbook.py "/path/to/2026 Standup Call Log.xlsx"
"""

from __future__ import annotations

import argparse
import json
import os
import urllib.parse
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook


DEFAULT_ORG_ID = "00000000-0000-0000-0000-000000000001"
FACILITY_NAMES = ["Homewood", "Oakridge", "Rising Oaks", "Plantation", "Grande Cypress"]
METRIC_MAP = {
    "Goal": "ar_goal_cents",
    "Current AR": "current_ar_cents",
    "Current  Total Census": "current_total_census",
    "Average Rent": "average_rent_cents",
    "Uncollected AR Total": "uncollected_ar_total_cents",
    "SP Female Beds Open": "sp_female_beds_open",
    "SP Male Beds Open": "sp_male_beds_open",
    "SP Male or Female Beds Open": "sp_flexible_beds_open",
    "Private Beds Open": "private_beds_open",
    "Total Beds open": "total_beds_open",
    "Admissions Expected": "admissions_expected",
    "Total at the Hospital & Rehab": "hospital_and_rehab_total",
    "Expected Discharges": "expected_discharges",
    "Call Outs Last Week": "callouts_last_week",
    "Terminations Last Week": "terminations_last_week",
    "Current Open Positions": "current_open_positions",
    "Overtime": "overtime_hours",
    "Tours Expected": "tours_expected",
    "Activities on the calendar to be completed by Home Health Providers": "provider_activities_expected",
    "Outreach & Engagements (Providers, Facilities, Events)": "outreach_engagements",
}


@dataclass
class Context:
    url: str
    key: str
    organization_id: str
    facility_name_to_id: dict[str, str]
    import_job_id: str


def env(name: str) -> str:
    value = os.environ.get(name, "").strip()
    if not value:
        raise RuntimeError(f"Missing environment variable: {name}")
    return value


def rest_request(method: str, path: str, *, payload: Any | None = None, query: dict[str, str] | None = None, headers: dict[str, str] | None = None) -> Any:
    base = env("NEXT_PUBLIC_SUPABASE_URL")
    key = env("SUPABASE_SERVICE_ROLE_KEY")
    params = urllib.parse.urlencode(query or {})
    url = f"{base.rstrip('/')}/rest/v1/{path}"
    if params:
      url = f"{url}?{params}"

    req_headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
    }
    if headers:
        req_headers.update(headers)
    data = None if payload is None else json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(url, method=method, data=data, headers=req_headers)
    with urllib.request.urlopen(req) as res:
        body = res.read().decode("utf-8")
        if not body:
            return None
        return json.loads(body)


def normalize_label(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        return s or None
    return None


def to_week_of(value: Any) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def parse_numeric(value: Any) -> float | None:
    if value in (None, "", " "):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"#DIV/0!", "—"}:
            return None
        try:
            return float(stripped)
        except ValueError:
            return None
    return None


def get_or_create_import_job(organization_id: str, file_name: str) -> str:
    row = rest_request(
        "POST",
        "exec_standup_import_jobs",
        payload={
            "organization_id": organization_id,
            "source_file_name": file_name,
            "status": "running",
            "started_at": datetime.utcnow().isoformat(),
            "source_ref_json": {"import_mode": "workbook"},
        },
        headers={"Prefer": "return=representation"},
    )
    return row[0]["id"]


def update_import_job(import_job_id: str, **fields: Any) -> None:
    rest_request(
        "PATCH",
        "exec_standup_import_jobs",
        payload=fields,
        query={"id": f"eq.{import_job_id}"},
        headers={"Prefer": "return=minimal"},
    )


def fetch_facilities(organization_id: str) -> dict[str, str]:
    rows = rest_request(
        "GET",
        "facilities",
        query={
            "organization_id": f"eq.{organization_id}",
            "deleted_at": "is.null",
            "select": "id,name",
        },
    ) or []
    return {row["name"]: row["id"] for row in rows}


def upsert_snapshot(organization_id: str, week_of: str, import_job_id: str) -> str:
    existing = rest_request(
        "GET",
        "exec_standup_snapshots",
        query={
            "organization_id": f"eq.{organization_id}",
            "week_of": f"eq.{week_of}",
            "deleted_at": "is.null",
            "select": "id",
        },
    ) or []
    if existing:
        snapshot_id = existing[0]["id"]
        rest_request(
            "DELETE",
            "exec_standup_snapshot_metrics",
            query={"snapshot_id": f"eq.{snapshot_id}"},
            headers={"Prefer": "return=minimal"},
        )
        rest_request(
            "PATCH",
            "exec_standup_snapshots",
            payload={
                "status": "published",
                "confidence_band": "low",
                "summary_json": {"import_job_id": import_job_id, "imported": True},
            },
            query={"id": f"eq.{snapshot_id}"},
            headers={"Prefer": "return=minimal"},
        )
        return snapshot_id

    created = rest_request(
        "POST",
        "exec_standup_snapshots",
        payload={
            "organization_id": organization_id,
            "week_of": week_of,
            "status": "published",
            "generated_at": datetime.utcnow().isoformat(),
            "published_at": datetime.utcnow().isoformat(),
            "published_version": 1,
            "confidence_band": "low",
            "completeness_pct": 0,
            "summary_json": {"import_job_id": import_job_id, "imported": True},
        },
        headers={"Prefer": "return=representation"},
    )
    return created[0]["id"]


def import_workbook(path: str, organization_id: str) -> None:
    wb = load_workbook(path, data_only=True)
    import_job_id = get_or_create_import_job(organization_id, os.path.basename(path))
    facility_map = fetch_facilities(organization_id)
    ctx = Context(
        url=env("NEXT_PUBLIC_SUPABASE_URL"),
        key=env("SUPABASE_SERVICE_ROLE_KEY"),
        organization_id=organization_id,
        facility_name_to_id=facility_map,
        import_job_id=import_job_id,
    )

    imported_weeks = 0
    imported_metrics = 0

    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            i = 0
            while i < len(rows):
                week_of = to_week_of(rows[i][0] if rows[i] else None)
                if not week_of:
                    i += 1
                    continue

                snapshot_id = upsert_snapshot(organization_id, week_of, import_job_id)
                metric_rows = []
                j = i + 2
                while j < len(rows):
                    first = rows[j][0] if rows[j] else None
                    if to_week_of(first):
                        break
                    label = normalize_label(first)
                    if label and label in METRIC_MAP:
                        metric_key = METRIC_MAP[label]
                        for offset, facility_name in enumerate(FACILITY_NAMES, start=1):
                            facility_id = ctx.facility_name_to_id.get(facility_name)
                            if not facility_id:
                                continue
                            raw_value = rows[j][offset] if offset < len(rows[j]) else None
                            metric_rows.append({
                                "snapshot_id": snapshot_id,
                                "organization_id": organization_id,
                                "facility_id": facility_id,
                                "section_key": infer_section(metric_key),
                                "metric_key": metric_key,
                                "metric_label": label,
                                "value_numeric": parse_numeric(raw_value),
                                "value_text": None if parse_numeric(raw_value) is not None else (str(raw_value).strip() if raw_value not in (None, "", " ") else None),
                                "value_currency_code": "USD",
                                "source_mode": "manual",
                                "confidence_band": "low",
                                "totals_included": False,
                                "freshness_at": None,
                                "source_ref_json": [{"sheet": ws.title, "row": j + 1, "import_job_id": import_job_id}],
                                "override_note": "Imported from legacy standup workbook.",
                            })
                        total_value = rows[j][6] if len(rows[j]) > 6 else None
                        metric_rows.append({
                            "snapshot_id": snapshot_id,
                            "organization_id": organization_id,
                            "facility_id": None,
                            "section_key": infer_section(metric_key),
                            "metric_key": metric_key,
                            "metric_label": label,
                            "value_numeric": parse_numeric(total_value),
                            "value_text": None if parse_numeric(total_value) is not None else (str(total_value).strip() if total_value not in (None, "", " ") else None),
                            "value_currency_code": "USD",
                            "source_mode": "manual",
                            "confidence_band": "low",
                            "totals_included": True,
                            "freshness_at": None,
                            "source_ref_json": [{"sheet": ws.title, "row": j + 1, "import_job_id": import_job_id}],
                            "override_note": "Imported from legacy standup workbook.",
                        })
                    j += 1

                if metric_rows:
                    rest_request(
                        "POST",
                        "exec_standup_snapshot_metrics",
                        payload=metric_rows,
                        headers={"Prefer": "return=minimal"},
                    )
                    completeness = round((sum(1 for row in metric_rows if row["value_numeric"] is not None or row["value_text"]) / len(metric_rows)) * 100, 2)
                    rest_request(
                        "PATCH",
                        "exec_standup_snapshots",
                        payload={"completeness_pct": completeness},
                        query={"id": f"eq.{snapshot_id}"},
                        headers={"Prefer": "return=minimal"},
                    )
                    imported_weeks += 1
                    imported_metrics += len(metric_rows)
                i = j

        update_import_job(
            import_job_id,
            status="completed",
            imported_week_count=imported_weeks,
            imported_metric_count=imported_metrics,
            finished_at=datetime.utcnow().isoformat(),
            result_json={"imported_weeks": imported_weeks, "imported_metrics": imported_metrics},
        )
        print(f"Imported {imported_weeks} weeks / {imported_metrics} metric rows.")
    except Exception as exc:  # noqa: BLE001
        update_import_job(
            import_job_id,
            status="failed",
            error_text=str(exc),
            finished_at=datetime.utcnow().isoformat(),
        )
        raise


def infer_section(metric_key: str) -> str:
    if metric_key in {"ar_goal_cents", "current_ar_cents", "current_total_census", "average_rent_cents", "uncollected_ar_total_cents"}:
        return "ar_census"
    if metric_key in {"sp_female_beds_open", "sp_male_beds_open", "sp_flexible_beds_open", "private_beds_open", "total_beds_open"}:
        return "bed_availability"
    if metric_key in {"admissions_expected"}:
        return "admissions"
    if metric_key in {"hospital_and_rehab_total", "expected_discharges"}:
        return "risk_management"
    if metric_key in {"callouts_last_week", "terminations_last_week", "current_open_positions", "overtime_hours"}:
        return "staffing"
    return "marketing"


def main() -> None:
    parser = argparse.ArgumentParser(description="Import the legacy standup workbook into exec_standup_* tables.")
    parser.add_argument("path", help="Absolute or relative path to the workbook .xlsx file.")
    parser.add_argument(
        "--org-id",
        default=os.environ.get("HAVEN_ORGANIZATION_ID", DEFAULT_ORG_ID),
        help="Target organization UUID. Defaults to HAVEN_ORGANIZATION_ID or the demo organization id.",
    )
    args = parser.parse_args()
    import_workbook(args.path, args.org_id)


if __name__ == "__main__":
    main()
